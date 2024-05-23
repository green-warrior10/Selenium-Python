# First we need to install library openpyxl, for this in console introduce -> pip install openpyxl
import openpyxl


class Funcion_Excel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file,path,sheetName):
        Workbook=openpyxl.load_workbook(path)
        sheet= Workbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file,sheetName):
        Workbook= openpyxl.load_workbook(file)
        sheet=  Workbook[sheetName]
        return (sheet.max_column)

    def readData(file,path,sheetName,no_rown,no_column):
        Workbook = openpyxl.load_workbook(path)
        sheet =  Workbook[sheetName]
        return sheet.cell(row=no_rown, column=no_column).value

    def writeData(file,path,sheetName,no_row,no_column,data):
        Workbook = openpyxl.load_workbook(path)
        sheet =  Workbook[sheetName]
        sheet.cell(row=no_row, column=no_column).value = data
        Workbook.save(path)

